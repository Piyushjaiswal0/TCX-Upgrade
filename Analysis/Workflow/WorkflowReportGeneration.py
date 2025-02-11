import tkinter as tk
from tkinter import filedialog, messagebox
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import xml.etree.ElementTree as ET
from collections import Counter

# Function to load WorkflowTemplate Object types from Excel
def load_workflow_template_object_types(excel_file_path):
    df = pd.read_excel(excel_file_path, sheet_name='WorkflowTemplateTypes')
    return df['WorkflowTemplate Object type'].dropna().unique()

# Function to load unsupported Workflow Handlers from CSV
def load_not_supported_workflow_handlers(csv_file_path):
    df = pd.read_csv(csv_file_path)
    return set(df['WorkflowHandler'].dropna())

# Function to parse XML and extract relevant data
def parse_xml(xml_file, workflow_template_object_types, not_supported_handlers):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    namespace = {'plm': 'http://www.plmxml.org/Schemas/PLMXMLSchema'}
    data_workflow_template = []
    data_workflow_handler = []

    for workflow_template_object_type in workflow_template_object_types:
        workflow_templates = root.findall(f'.//plm:WorkflowTemplate[@objectType="{workflow_template_object_type}"]', namespace)
        workflow_template_count = len(workflow_templates)
        if workflow_template_count > 0:
            data_workflow_template.append([xml_file, workflow_template_object_type, workflow_template_count])

    workflow_handlers = root.findall('.//plm:WorkflowHandler', namespace)
    handler_names = [handler.attrib.get('name') for handler in workflow_handlers]
    handler_count = Counter(handler_names)

    for handler_name, count in handler_count.items():
        if handler_name in not_supported_handlers:
            data_workflow_handler.append([xml_file, handler_name, count])

    return data_workflow_template, data_workflow_handler

# Function to save the data to Excel
def save_to_excel(data_workflow_template, data_workflow_handler, output_excel_path):
    df_workflow_template = pd.DataFrame(data_workflow_template, columns=["File Name with path", "WorkflowTemplate Object type", "Total WorkflowTemplate Count"])
    df_workflow_handler = pd.DataFrame(data_workflow_handler, columns=["File Name with path", "WorkflowHandler", "WorkflowTemplate Count"])

    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df_workflow_template.to_excel(writer, index=False, sheet_name='WorkflowTemplate')
        df_workflow_handler.to_excel(writer, index=False, sheet_name='WorkflowHandler')

    wb = load_workbook(output_excel_path)
    ws_workflow_handler = wb['WorkflowHandler']
    ws_workflow_template = wb['WorkflowTemplate']

    # Merge cells for both sheets using a helper function
    merge_cells_for_same_file_name(ws_workflow_handler, df_workflow_handler, 'WorkflowHandler')
    merge_cells_for_same_file_name(ws_workflow_template, df_workflow_template, 'WorkflowTemplate')

    wb.save(output_excel_path)

# Function to merge cells for same file name
def merge_cells_for_same_file_name(ws, data_frame, sheet_name):
    groups = {}
    for _, row in data_frame.iterrows():
        file_name = row['File Name with path']
        groups.setdefault(file_name, []).append((row['WorkflowTemplate Object type'] if sheet_name == 'WorkflowTemplate' else row['WorkflowHandler'], row['Total WorkflowTemplate Count'] if sheet_name == 'WorkflowTemplate' else row['WorkflowTemplate Count']))

    ws.delete_rows(2, ws.max_row)
    row = 2
    for file_name, entries in groups.items():
        ws[f"A{row}"] = file_name
        merge_start_row = row

        for entry in entries:
            ws[f"B{row}"] = entry[0]  # Object Type or Handler Name
            ws[f"C{row}"] = entry[1]  # Count
            row += 1

        if row - merge_start_row > 1:
            ws.merge_cells(f"A{merge_start_row}:A{row-1}")

        for i in range(merge_start_row, row):
            ws[f"A{i}"].alignment = Alignment(vertical='center')

# Function to run the process
def run_process():
    input_excel_file_path = input_excel_var.get()
    input_csv_file_path = input_csv_var.get()
    directory_path = directory_var.get()
    output_excel_path = output_excel_var.get()

    if not all([input_excel_file_path, input_csv_file_path, directory_path, output_excel_path]):
        messagebox.showerror("Error", "Please provide all the required paths.")
        return

    workflow_template_object_types = load_workflow_template_object_types(input_excel_file_path)
    not_supported_handlers = load_not_supported_workflow_handlers(input_csv_file_path)

    data_workflow_template = []
    data_workflow_handler = []

    for filename in os.listdir(directory_path):
        if filename.endswith('.xml'):
            xml_file_path = os.path.join(directory_path, filename)
            workflow_template, workflow_handler = parse_xml(xml_file_path, workflow_template_object_types, not_supported_handlers)
            data_workflow_template.extend(workflow_template)
            data_workflow_handler.extend(workflow_handler)

    save_to_excel(data_workflow_template, data_workflow_handler, output_excel_path)
    messagebox.showinfo("Success", f"Filtered Excel report generated: {output_excel_path}")

# Create the GUI window
root = tk.Tk()
root.title("Workflow Report Generator")

# Create the input and output fields
tk.Label(root, text="Input Excel File for WorkflowTemplate Object types").grid(row=0, column=0, padx=10, pady=5)
input_excel_var = tk.StringVar()
tk.Entry(root, textvariable=input_excel_var, width=50).grid(row=0, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=lambda: input_excel_var.set(filedialog.askopenfilename())).grid(row=0, column=2, padx=10, pady=5)

tk.Label(root, text="Workflow Handlers which are not supported in TCX").grid(row=1, column=0, padx=10, pady=5)
input_csv_var = tk.StringVar()
tk.Entry(root, textvariable=input_csv_var, width=50).grid(row=1, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=lambda: input_csv_var.set(filedialog.askopenfilename())).grid(row=1, column=2, padx=10, pady=5)

tk.Label(root, text="Repository Directory Path").grid(row=2, column=0, padx=10, pady=5)
directory_var = tk.StringVar()
tk.Entry(root, textvariable=directory_var, width=50).grid(row=2, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=lambda: directory_var.set(filedialog.askdirectory())).grid(row=2, column=2, padx=10, pady=5)

tk.Label(root, text="Enter Output File Dir").grid(row=3, column=0, padx=10, pady=5)
output_excel_var = tk.StringVar()
tk.Entry(root, textvariable=output_excel_var, width=50).grid(row=3, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=lambda: output_excel_var.set(filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")]))).grid(row=3, column=2, padx=10, pady=5)

# Run Button
tk.Button(root, text="Generate Report", command=run_process, width=20).grid(row=4, column=1, padx=10, pady=20)

# Start the GUI event loop
root.mainloop()
