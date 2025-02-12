from openpyxl.styles import Alignment

# Function to merge cells in a specified column with similar contents and apply alignment
def merge_cells_in_column(ws, column_index):

    current_value = None
    merge_start_row = None

    # Iterate through all rows in the specified column (skip header row)
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=column_index).value

        if cell_value == current_value:
            continue
        else:
            if merge_start_row is not None and row > merge_start_row + 1:
                # Merge cells in the range [merge_start_row, row - 1] for the specified column
                ws.merge_cells(start_row=merge_start_row, start_column=column_index, end_row=row - 1, end_column=column_index)
                # Apply center alignment both vertically
                for r in range(merge_start_row, row):
                    cell = ws.cell(row=r, column=column_index)
                    cell.alignment = Alignment(vertical='center')
            
            # Set new merge start point
            merge_start_row = row
            current_value = cell_value

    # Handle last merge if needed
    if merge_start_row is not None and ws.max_row > merge_start_row:
        ws.merge_cells(start_row=merge_start_row, start_column=column_index, end_row=ws.max_row, end_column=column_index)
        for r in range(merge_start_row, ws.max_row + 1):
            cell = ws.cell(row=r, column=column_index)
            cell.alignment = Alignment(vertical='center')
