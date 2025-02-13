import os
import pandas as pd
from collections import Counter
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from CommonFunctions.ExcelCommonFunctions import merge_cells_in_column

# Function to save the data to Excel with merged cells and alignment
def save_to_excel(data_workflow_handler, output_excel_path):
    # Save the extracted data to an Excel file using pandas
    df_workflow_handler = pd.DataFrame(data_workflow_handler, columns=["File Name with path", "Workflow Template Name", "WorkflowHandler", "WorkflowTemplate Count"])
    df_workflow_handler.to_excel(output_excel_path, index=False, sheet_name='WorkflowHandler')

    # Now use openpyxl to open the saved workbook and apply cell merging and alignment
    wb = load_workbook(output_excel_path)
    ws = wb['WorkflowHandler']

    # Merge cells in column A and B by calling the merge function
    merge_cells_in_column(ws, 1)
    merge_cells_in_column(ws, 2)

    # Save the workbook after modifications
    wb.save(output_excel_path)

def parse_xml(xml_file, workflow_template_object_types, not_supported_handlers):
    # Parse XML file and extract workflow handler data.
    tree = ET.parse(xml_file)
    root = tree.getroot()

    namespace = {'plm': 'http://www.plmxml.org/Schemas/PLMXMLSchema'}
    data_workflow_handler = []
    
    # Store handlers and their counts in a dictionary for faster lookup
    handler_count = Counter(handler.attrib.get('name') for handler in root.findall('.//plm:WorkflowHandler', namespace))
    
    # Find and process workflow templates for each object type
    for workflow_template_object_type in workflow_template_object_types:
        workflow_templates = root.findall(f'.//plm:WorkflowTemplate[@objectType="{workflow_template_object_type}"]', namespace)
        for workflow_template in workflow_templates:
            workflow_template_name = workflow_template.attrib.get('name')

            # Filter and append unsupported handlers to results
            for handler_name, count in handler_count.items():
                if handler_name in not_supported_handlers:
                    data_workflow_handler.append([xml_file, workflow_template_name, handler_name, count])

    return data_workflow_handler

def transform_to_property_structure(data_workflow_handler):
    transformed_data = {}
    
    # Process each entry in data_workflow_handler
    for entry in data_workflow_handler:
        # Generate a unique ID for each entry based on a combination of file name and template name
        property_id = len(transformed_data) + 1
        transformed_data[property_id] = {
            'File Name with path': entry[0],
            'Workflow Template Name': entry[1],
            'Workflow Handler': entry[2],
            'Workflow Template Count': entry[3]
        }

    return transformed_data

# Main function to generate the workflow report
def generate_workflow_report(workflow_template_object_types, not_supported_workflows, workflow_code_dir, output_dir):
    data_workflow_handler = []
    workspace_data_dict = {}

    for filename in os.listdir(workflow_code_dir):
        if filename.endswith('.xml'):
            xml_file_path = os.path.join(workflow_code_dir, filename)
            data_workflow_handler.extend(parse_xml(xml_file_path, workflow_template_object_types, not_supported_workflows))

    # Save the data to Excel
    output_location = os.path.join(output_dir, "OutputWorkflowHandlers.xlsx")
    save_to_excel(data_workflow_handler, output_location)

    # Transform the data into the desired structure
    workspace_data_dict = transform_to_property_structure(data_workflow_handler)
    print(f"Report Generated for \"Workflow Handlers\" has been saved to {output_location}\n")

    return workspace_data_dict
